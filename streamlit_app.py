import streamlit as st
import pandas as pd
import numpy as np
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from datetime import datetime


import streamlit as st
import pandas as pd
import io
from datetime import datetime

# Add this at the beginning of your existing code, after the imports
def main():
    st.set_page_config(
        page_title="Cooling Circuit Calculator",
        page_icon="🔧",
        layout="wide"
    )
    
    st.title("🔧 Cooling Circuit Calculator")
    st.markdown("---")
    
    # Initialize session state for results
    if 'results' not in st.session_state:
        st.session_state.results = None
    if 'inputs' not in st.session_state:
        st.session_state.inputs = None
    
    # Create two columns for input and output
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.header("📝 Input Parameters")
        
        # Create input form
        with st.form("cooling_inputs"):
            st.subheader("Circuit Configuration")
            NK = st.number_input("Number of cooling circuits (NK)", min_value=1, max_value=9, value=1, step=1)
            NB = st.number_input("Cooling layer (NB)", min_value=2, max_value=9, value=2, step=1)
            
            st.subheader("Iron Dimensions")
            HSP = st.number_input("Iron height (HSP) [mm]", min_value=0.0, value=100.0, step=0.1)
            ASP = st.number_input("Iron length (ASP) [mm]", min_value=0.0, value=150.0, step=0.1)
            min_bending_radius = st.number_input("Min bending radius [mm]", min_value=0.0, value=5.0, step=0.1)
            
            st.subheader("Electrical Parameters")
            I = st.number_input("Current (I) [A]", min_value=0.0, value=1000.0, step=1.0)
            NL = st.number_input("Number of windings per layer (NL)", min_value=1, value=10, step=1)
            
            st.subheader("Copper Profile")
            BCU = st.number_input("Copper profile breadth (BCU) [mm]", min_value=0.0, value=15.0, step=0.1)
            HCU = st.number_input("Copper profile height (HCU) [mm]", min_value=0.0, value=8.0, step=0.1)
            TCU = st.number_input("Thickness (TCU) [mm]", min_value=0.0, value=2.0, step=0.1)
            RICU = st.number_input("Inner diameter (RICU) [mm]", min_value=0.0, value=3.0, step=0.1)
            RYCU = st.number_input("Outer diameter (RYCU) [mm]", min_value=0.0, value=6.0, step=0.1)
            
            st.subheader("Temperature Parameters")
            Tin = st.number_input("Inlet temperature (Tin) [°C]", value=40.0, step=0.1)
            Tmax = st.number_input("Maximum temperature (Tmax) [°C]", value=80.0, step=0.1)
            TCISO = st.selectbox("TCISO value", options=[2.4, 2.8, 3.2], index=0)
            
            # Submit button
            submitted = st.form_submit_button("🚀 Calculate", use_container_width=True)
            
            if submitted:
                # Prepare inputs dictionary
                inputs = {
                    'NK': NK,
                    'NB': NB,
                    'HSP': HSP,
                    'ASP': ASP,
                    'min_bending_radius': min_bending_radius,
                    'I': I,
                    'NL': NL,
                    'BCU': BCU,
                    'HCU': HCU,
                    'TCU': TCU,
                    'RICU': RICU,
                    'RYCU': RYCU,
                    'Tin': Tin,
                    'Tmax': Tmax,
                    'TCISO': TCISO
                }
                
                # Validate inputs
                try:
                    calculator = CoolingCircuitCalculator()
                    validated_inputs = calculator.validate_inputs(inputs)
                    st.session_state.inputs = validated_inputs
                    
                    # Calculate results
                    with st.spinner('Calculating...'):
                        results = calculator.calculate(validated_inputs)
                        st.session_state.results = results
                    
                    st.success("✅ Calculation completed successfully!")
                    
                except Exception as e:
                    st.error(f"❌ Error in calculation: {str(e)}")
    
    with col2:
        st.header("📊 Results")
        
        if st.session_state.results is not None:
            results = st.session_state.results
            inputs = st.session_state.inputs
            
            # Display results based on type
            if isinstance(results, dict) and 'columns' in results:
                # Multi-column results
                st.subheader(f"Results for {inputs['NK']} Cooling Circuits")
                
                # Create tabs for different views
                tab1, tab2, tab3 = st.tabs(["📋 Summary", "📈 Detailed Results", "📊 Totals"])
                
                with tab1:
                    # Summary table
                    summary_data = []
                    for i, col_result in enumerate(results['columns']):
                        summary_data.append({
                            'Column': f"Column {i+1}",
                            'Copper Loss (kW)': round(col_result['Pcu'], 4),
                            'Water Flow (L/m)': round(col_result['Q'], 4),
                            'Water Velocity (m/s)': round(col_result['Vh2O'], 4),
                            'Pressure Drop (kPa)': round(col_result['dp_kpa'], 4),
                            'Temperature Rise (°C)': round(col_result['dT'], 4)
                        })
                    
                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(summary_df, use_container_width=True)
                
                with tab2:
                    # Detailed results for each column
                    selected_column = st.selectbox("Select Column", 
                                                 [f"Column {i+1}" for i in range(len(results['columns']))])
                    
                    col_idx = int(selected_column.split()[-1]) - 1
                    col_result = results['columns'][col_idx]
                    
                    # Display detailed results
                    detailed_data = {
                        'Parameter': ['Cooling Area', 'Copper Area', 'Hydraulic Diameter', 'Copper Length', 
                                    'Copper Weight', 'Conductivity', 'Copper Resistivity', 'Water Flow',
                                    'Water Velocity', 'Copper Loss', 'Temperature Rise', 'Reynolds Number',
                                    'Friction Factor', 'Pressure Drop'],
                        'Symbol': ['Akyl', 'Acu', 'Dh', 'Lcu', 'Mcu', 'σ', 'Recu', 'Q', 
                                 'Vh2O', 'Pcu', 'dT', 'Re', 'f', 'Δp'],
                        'Value': [col_result['Akyl'], col_result['Acu'], col_result['Dh'], col_result['Lcu'],
                                col_result['Mcu'], col_result['Conductivity'], col_result['Recu'], col_result['Q'],
                                col_result['Vh2O'], col_result['Pcu'], col_result['dT'], col_result['Re'],
                                col_result['f'], col_result['dp_kpa']],
                        'Unit': ['mm²', 'mm²', 'mm', 'mm', 'kg', 'A/mm²', 'Ωmm²/m', 'L/m',
                               'm/s', 'kW', '°C', '', '', 'kPa']
                    }
                    
                    detailed_df = pd.DataFrame(detailed_data)
                    detailed_df['Value'] = detailed_df['Value'].apply(lambda x: round(x, 4) if isinstance(x, float) else x)
                    st.dataframe(detailed_df, use_container_width=True)
                
                with tab3:
                    # Totals and averages
                    col_a, col_b = st.columns(2)
                    
                    with col_a:
                        st.metric("Total Copper Loss", f"{results['total_pcu']:.4f} kW")
                        st.metric("Total Water Flow", f"{results['total_q']:.4f} L/m")
                        st.metric("Total Copper Length", f"{results['total_lcu']:.4f} mm")
                    
                    with col_b:
                        st.metric("Average Water Velocity", f"{results['averages']['Vh2O']:.4f} m/s")
                        st.metric("Average Pressure Drop", f"{results['averages']['dp_kpa']:.4f} kPa")
                        st.metric("Average Reynolds Number", f"{results['averages']['Re']:.0f}")
            
            else:
                # Single column results
                st.subheader("Single Column Results")
                
                # Key metrics
                col_a, col_b, col_c = st.columns(3)
                with col_a:
                    st.metric("Copper Loss", f"{results['Pcu']:.4f} kW")
                    st.metric("Water Flow", f"{results['Q']:.4f} L/m")
                
                with col_b:
                    st.metric("Water Velocity", f"{results['Vh2O']:.4f} m/s")
                    st.metric("Pressure Drop", f"{results['dp_kpa']:.4f} kPa")
                
                with col_c:
                    st.metric("Temperature Rise", f"{results['dT']:.4f} °C")
                    st.metric("Reynolds Number", f"{results['Re']:.0f}")
                
                # Detailed results table
                st.subheader("Detailed Results")
                detailed_data = {
                    'Parameter': ['Cooling Area', 'Copper Area', 'Hydraulic Diameter', 'Copper Length', 
                                'Copper Weight', 'Conductivity', 'Copper Resistivity', 'Water Flow',
                                'Water Velocity', 'Copper Loss', 'Temperature Rise', 'Reynolds Number',
                                'Friction Factor', 'Pressure Drop'],
                    'Symbol': ['Akyl', 'Acu', 'Dh', 'Lcu', 'Mcu', 'σ', 'Recu', 'Q', 
                             'Vh2O', 'Pcu', 'dT', 'Re', 'f', 'Δp'],
                    'Value': [results['Akyl'], results['Acu'], results['Dh'], results['Lcu'],
                            results['Mcu'], results['Conductivity'], results['Recu'], results['Q'],
                            results['Vh2O'], results['Pcu'], results['dT'], results['Re'],
                            results['f'], results['dp_kpa']],
                    'Unit': ['mm²', 'mm²', 'mm', 'mm', 'kg', 'A/mm²', 'Ωmm²/m', 'L/m',
                           'm/s', 'kW', '°C', '', '', 'kPa']
                }
                
                detailed_df = pd.DataFrame(detailed_data)
                detailed_df['Value'] = detailed_df['Value'].apply(lambda x: round(x, 4) if isinstance(x, float) else x)
                st.dataframe(detailed_df, use_container_width=True)
            
            # Download button
            if st.button("📥 Download Excel Report", use_container_width=True):
                # Create Excel file in memory
                output = io.BytesIO()
                calculator = CoolingCircuitCalculator()
                
                # Save to temporary file first
                temp_filename = f"cooling_circuit_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                calculator.save_to_excel(inputs, results, temp_filename)
                
                # Read the file and create download
                with open(temp_filename, 'rb') as f:
                    excel_data = f.read()
                
                st.download_button(
                    label="Click to Download Excel File",
                    data=excel_data,
                    file_name=temp_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Clean up temporary file
                import os
                if os.path.exists(temp_filename):
                    os.remove(temp_filename)
        
        else:
            st.info("👈 Please enter input parameters and click 'Calculate' to see results")
    
    # Footer
    st.markdown("---")
    st.markdown("**Cooling Circuit Calculator** - Built with Streamlit")

class CoolingCircuitCalculator:
    def __init__(self):
        # Constants as provided in the notes
        self.PI = 3.141593
        self.g = 10  # gravity constant
        self.recu20 = 0.017241  # resistivity of copper (ohm.mm²/m)
        self.tkoeff = 0.00393  # temperature coefficient
        self.dens = 1000  # density (water presumably)
        self.cp = 4186  # specific heat capacity
        self.ticu = 0.45  # thermal conductivity of copper
        self.tib = 0.4  # thermal conductivity of insulation
        self.v_40c = 0.458  # kinematic viscosity at 40°C
        self.e = 0.0015  # roughness factor
        self.tciso_values = [2.4, 2.8, 3.2]  # possible TCISO values
    
    def calculate_friction_factor_with_logic(self, Re, Dh, dp_kpa):
        """
        Calculate friction factor with logic-based adjustment based on pressure drop.
        
        Args:
            Re: Reynolds number
            Dh: Hydraulic diameter
            dp_kpa: Pressure drop in kPa
            
        Returns:
            Adjusted friction factor
        """
        # Initial friction factor calculation
        if Re <= 0:
            f = 0.02
        elif Re < 2300:
            f = 64 / Re
        else:
            relative_roughness = self.e / Dh
            f = 0.25 / (math.log10(relative_roughness/3.7 + 5.74/Re**0.9))**2
            
            try:
                for i in range(3):
                    f_inv = -2 * math.log10(relative_roughness/3.7 + 2.51/(Re*math.sqrt(f)))
                    f_new = 1 / (f_inv * f_inv)
                    f = f_new
            except (ValueError, ZeroDivisionError):
                pass
        
        # Logic-based adjustment based on pressure drop
        if dp_kpa < 400:
            # Logic 1: Add 0.008 to friction factor
            f_adjusted = f + 0.008
        elif 400 <= dp_kpa <= 450:
            # Logic 2: Add 0.004 to friction factor  
            f_adjusted = f + 0.004
        else:  # dp_kpa > 450
            # Logic 3: Keep original friction factor (no adjustment)
            f_adjusted = f
        
        return f_adjusted
    
    def calculate_single_column(self, inputs, column_num=1):
        """
        Calculate results for a single cooling column.
        
        Args:
            inputs: Dictionary with input parameters
            column_num: Column number (1-based)
            
        Returns:
            Dictionary with calculated results for this column
        """
        # Extract inputs for easier reference
        NK = inputs['NK']
        NB = inputs['NB']
        HSP = inputs['HSP']
        ASP = inputs['ASP']
        min_bending_radius = inputs['min_bending_radius']
        I = inputs['I']
        NL = inputs['NL']
        BCU = inputs['BCU']
        HCU = inputs['HCU']
        TCU = inputs['TCU']
        RICU = inputs['RICU']
        RYCU = inputs['RYCU']
        Tin = inputs['Tin']
        Tmax = inputs['Tmax']
        TCISO = inputs['TCISO']
        
        # 1. Cooling area (Akyl) calculation 
        b = BCU - 2 * TCU
        h = HCU - 2 * TCU
        r = RICU
        
    
        # Acu = (BCU × HCU - 4 × (1 - π/4) × RYCU²) - ((BCU - 2×RICU) × (HCU - 2×RICU) - 4 × (1 - π/4) × RICU²)
        Acu = (BCU * HCU - 4 * (1 - self.PI/4) * RYCU**2) - ((BCU - 2*TCU) * (HCU - 2*TCU) - 4 * (1 - self.PI/4) * RICU**2)
        
        # Cooling area calculation
        Akyl = (b - 2*r) * h + 2 * r * (h - 2*r) + self.PI * r * r
        
        # 2. Hydraulic diameter (Dh) calculation
        area = 2 * (b - 2*r) + 2 * (h - 2*r) + 2 * r * self.PI
        Dh = 4 * Akyl / area
        
        # 3. Copper length (Lcu) calculation for this specific column
        ltot = 0
        oraka = 2 * (HSP - 2 * RICU) + 2 * (ASP - 2 * RICU)

        # Calculate radius for this specific column (column_num)
        radie = RICU + (column_num - 0.5) * (BCU + 2 * TCU) + (column_num - 1) * self.tib + TCISO



        # For this column, calculate length for all layers (NB)
        for layer in range(1, NB + 1):

            ltot += oraka + 2 * radie * self.PI

  

        Lcu = ltot * NL*15/100
    
                
        # 4. Copper weight (Mcu) calculation
        Mcu = self.dens * Lcu * Acu / 10**9
        
        # 5. Conductivity calculation
        conductivity = I / Acu
        
        # 6. Copper resistivity (Recu) calculation
        dT = Tmax - Tin
        cures = self.recu20 * (1 + self.tkoeff * (dT/2 + Tin - 20))
        
        # 7. Copper loss (Pcu) calculation - removed 10^6 factor as per notes

        Pcu = cures * Lcu * I**2 / Acu / 10**6
    
        # 8. Water flow (Q) calculation
        Q = 60 * (Pcu) * 1000 / (dT * self.cp)
        
        # 9. Water velocity (Vh2O) calculation
        Vh2O = Q / (60 * (Akyl / 10**6) * self.dens)
        
        # 10. Reynolds number calculation (Modified as per image)
        # Convert diameter to meters first
        Dh_m = Dh / 1000  # Convert from mm to meters

        # Calculate Reynolds number: Re = (v × D) / ν
        Re = (Vh2O * Dh_m) / (self.v_40c * 10**-6)  # v_40c is in mm²/s, convert to m²/s
        
        # 11. Initial friction factor calculation (without logic adjustment)
        if Re <= 0:
            print("Warning: Reynolds number is zero or negative. Using default friction factor.")
            f_initial = 0.02
        elif Re < 2300:
            f_initial = 64 / Re
        else:
            relative_roughness = self.e / Dh
            f_initial = 0.25 / (math.log10(relative_roughness/3.7 + 5.74/Re**0.9))**2
            
            try:
                for i in range(3):
                    f_inv = -2 * math.log10(relative_roughness/3.7 + 2.51/(Re*math.sqrt(f_initial)))
                    f_new = 1 / (f_inv * f_inv)
                    f_initial = f_new
            except (ValueError, ZeroDivisionError):
                pass
        
        # 12. Initial pressure drop calculation
        dp_initial = f_initial * Lcu / Dh * self.dens * Vh2O**2 / 2
        dp_kpa_initial = dp_initial / 1000
        
        # 13. Apply logic-based friction factor adjustment
        f = self.calculate_friction_factor_with_logic(Re, Dh, dp_kpa_initial)
        
        # 14. Final pressure drop calculation with adjusted friction factor
        dp = f * Lcu / Dh * self.dens * Vh2O**2 / 2
        dp_kpa = dp / 1000
        
        # 15. Change in temperature (dT) for this column
        dT_column = Pcu * 1000 * 60 / (self.cp * Q)
        
        return {
            'column': column_num,
            'Akyl': Akyl,
            'Acu': Acu,
            'Dh': Dh,
            'Lcu': Lcu,
            'Mcu': Mcu,
            'Conductivity': conductivity,
            'Recu': cures,
            'Q': Q,
            'Vh2O': Vh2O,
            'Pcu': Pcu,
            'dT': dT_column,
            'Re': Re,
            'f': f,
            'f_initial': f_initial,  # Added for debugging/comparison
            'dp_kpa': dp_kpa,
            'dp_kpa_initial': dp_kpa_initial,  # Added for debugging/comparison
            'Tmax_calc': dT_column + Tin
        }
    
    def calculate_multi_column(self, inputs):
        """
        Calculate results for multiple cooling columns with backward calculation for water velocity.
        
        Args:
            inputs: Dictionary with input parameters
            
        Returns:
            Dictionary with results for all columns and averages
        """
        NK = inputs['NK']
        all_results = []
        
        # Calculate first column normally
        column_1_results = self.calculate_single_column(inputs, 1)
        all_results.append(column_1_results)
        
        # For subsequent columns, use backward calculation as per notes
        for col in range(2, NK + 1):
            prev_results = all_results[-1]
            
            # Backward calculation for water velocity (from notes)
            # VH2O(2) = VH2O(1) * sqrt(LCU(1) / LCU(2))
            current_results = self.calculate_single_column(inputs, col)
            
            # Adjust water velocity based on previous column
            Vh2O_adjusted = prev_results['Vh2O'] * math.sqrt(prev_results['Lcu'] / current_results['Lcu'])
            current_results['Vh2O'] = Vh2O_adjusted
            
            # Recalculate water flow for this column
            current_results['Q'] = 60 * Vh2O_adjusted * (current_results['Akyl'] / 10**6) * self.dens
            
            # Recalculate Reynolds number (Modified as per image)
            Dh_m = current_results['Dh'] / 1000  # Convert from mm to meters
            Re = (Vh2O_adjusted * Dh_m) / (self.v_40c * 10**-6)  # v_40c is in mm²/s, convert to m²/s
            
            # Recalculate initial friction factor
            if Re <= 0:
                f_initial = 0.02
            elif Re < 2300:
                f_initial = 64 / Re
            else:
                relative_roughness = self.e / current_results['Dh']
                f_initial = 0.25 / (math.log10(relative_roughness/3.7 + 5.74/Re**0.9))**2
                
                try:
                    for i in range(3):
                        f_inv = -2 * math.log10(relative_roughness/3.7 + 2.51/(Re*math.sqrt(f_initial)))
                        f_new = 1 / (f_inv * f_inv)
                        f_initial = f_new
                except (ValueError, ZeroDivisionError):
                    pass
            
            # Calculate initial pressure drop
            dp_initial = f_initial * current_results['Lcu'] / current_results['Dh'] * self.dens * Vh2O_adjusted**2 / 2
            dp_kpa_initial = dp_initial / 1000
            
            # Apply logic-based friction factor adjustment
            f = self.calculate_friction_factor_with_logic(Re, current_results['Dh'], dp_kpa_initial)
            current_results['f'] = f
            current_results['f_initial'] = f_initial
            current_results['dp_kpa_initial'] = dp_kpa_initial
            
            # Recalculate final pressure drop with adjusted friction factor
            dp = f * current_results['Lcu'] / current_results['Dh'] * self.dens * Vh2O_adjusted**2 / 2
            current_results['dp_kpa'] = dp / 1000
            
            all_results.append(current_results)
        
        # Calculate averages
        averages = self.calculate_averages(all_results)
        
        return {
            'columns': all_results,
            'averages': averages,
            'total_pcu': sum([col['Pcu'] for col in all_results]),
            'total_q': sum([col['Q'] for col in all_results]),
            'total_lcu': sum([col['Lcu'] for col in all_results])
        }
    
    def calculate_averages(self, all_results):
        """Calculate average values across all columns."""
        if not all_results:
            return {}
        
        n_columns = len(all_results)
        averages = {}
        
        # Parameters to average
        params_to_average = ['Akyl', 'Acu', 'Dh', 'Lcu', 'Mcu', 'Conductivity', 
                           'Recu', 'Q', 'Vh2O', 'Pcu', 'dT', 'Re', 'f', 'dp_kpa']
        
        for param in params_to_average:
            averages[param] = sum([col[param] for col in all_results]) / n_columns
        
        return averages
    
    def calculate(self, inputs):
        """
        Main calculation method that handles both single and multi-column calculations.
        """
        if inputs['NK'] == 1:
            return self.calculate_single_column(inputs, 1)
        else:
            return self.calculate_multi_column(inputs)
    
    def save_to_excel(self, inputs, results, filename="cooling_circuit_inputs.xlsm"):
        """
        Save inputs and results to an Excel file with improved formatting for multi-column results.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cooling Circuit Calculation"
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Write title
        ws['A1'] = "Cooling Circuit Calculation Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Write inputs section
        ws['A3'] = "INPUTS:"
        ws['A3'].font = header_font
        
        row = 4
        for key, value in inputs.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Write outputs section
        ws[f'A{row+1}'] = "OUTPUTS:"
        ws[f'A{row+1}'].font = header_font
        
        output_row = row + 2
        
        # Check if multi-column results
        if isinstance(results, dict) and 'columns' in results:
            # Multi-column results
            # Headers
            headers = ["Parameter", "Symbol", "Unit"] + [f"Column {i+1}" for i in range(len(results['columns']))] + ["Average"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=output_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Output data
            output_data = [
                ["Cooling Area", "Akyl", "mm²"],
                ["Copper Area", "Acu", "mm²"],
                ["Hydraulic Diameter", "Dh", "mm"],
                ["Copper Length", "Lcu", "mm"],
                ["Copper Weight", "Mcu", "kg"],
                ["Conductivity", "σ", "A/mm²"],
                ["Copper Resistivity", "Recu", "Ωmm²/m"],
                ["Water Flow", "Q", "L/m"],
                ["Water Velocity", "Vh2O", "m/s"],
                ["Copper Loss", "Pcu", "kW"],
                ["Change in Temperature", "dT", "°C"],
                ["Reynolds Number", "Re", ""],
                ["Friction Factor (Adjusted)", "f", ""],
                ["Pressure Drop (Adjusted)", "Δp", "kPa"],
                ["Friction Factor (Initial)", "f_initial", ""],
                ["Pressure Drop (Initial)", "Δp_initial", "kPa"]
            ]
            
            param_keys = ['Akyl', 'Acu', 'Dh', 'Lcu', 'Mcu', 'Conductivity', 'Recu', 
                         'Q', 'Vh2O', 'Pcu', 'dT', 'Re', 'f', 'dp_kpa', 'f_initial', 'dp_kpa_initial']
            
            for r, (row_data, param_key) in enumerate(zip(output_data, param_keys), start=output_row+1):
                # Parameter info
                for c, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = cell_value
                    cell.border = border
                
                # Column values
                for col_idx, col_result in enumerate(results['columns']):
                    cell = ws.cell(row=r, column=4 + col_idx)
                    value = col_result.get(param_key, 'N/A')
                    cell.value = round(value, 4) if isinstance(value, float) else value
                    cell.border = border
                
                # Average value (only for main parameters, not initial values)
                if param_key in ['Akyl', 'Acu', 'Dh', 'Lcu', 'Mcu', 'Conductivity', 'Recu', 
                               'Q', 'Vh2O', 'Pcu', 'dT', 'Re', 'f', 'dp_kpa']:
                    avg_cell = ws.cell(row=r, column=4 + len(results['columns']))
                    avg_value = results['averages'][param_key]
                    avg_cell.value = round(avg_value, 4) if isinstance(avg_value, float) else avg_value
                    avg_cell.border = border
            
            # Add totals row
            totals_row = output_row + len(output_data) + 2
            ws[f'A{totals_row}'] = "TOTALS:"
            ws[f'A{totals_row}'].font = header_font
            
            ws[f'A{totals_row+1}'] = "Total Copper Loss (PCU)"
            ws[f'B{totals_row+1}'] = round(results['total_pcu'], 4)
            ws[f'A{totals_row+2}'] = "Total Water Flow (Q)"
            ws[f'B{totals_row+2}'] = round(results['total_q'], 4)
            ws[f'A{totals_row+3}'] = "Total Copper Length (LCU)"
            ws[f'B{totals_row+3}'] = round(results['total_lcu'], 4)
        
        else:
            # Single column results (original format)
            headers = ["Parameter", "Symbol", "Value", "Unit"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=output_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            output_data = [
                ["Cooling Area", "Akyl", results['Akyl'], "mm²"],
                ["Copper Area", "Acu", results['Acu'], "mm²"],
                ["Hydraulic Diameter", "Dh", results['Dh'], "mm"],
                ["Copper Length", "Lcu", results['Lcu'], "mm"],
                ["Copper Weight", "Mcu", results['Mcu'], "kg"],
                ["Conductivity", "σ", results['Conductivity'], "A/mm²"],
                ["Copper Resistivity", "Recu", results['Recu'], "Ωmm²/m"],
                ["Water Flow", "Q", results['Q'], "L/m"],
                ["Water Velocity", "Vh2O", results['Vh2O'], "m/s"],
                ["Copper Loss", "Pcu", results['Pcu'], "kW"],
                ["Change in Temperature", "dT", results['dT'], "°C"],
                ["Reynolds Number", "Re", results['Re'], ""],
                ["Friction Factor (Adjusted)", "f", results['f'], ""],
                ["Pressure Drop (Adjusted)", "Δp", results['dp_kpa'], "kPa"],
                ["Friction Factor (Initial)", "f_initial", results.get('f_initial', 'N/A'), ""],
                ["Pressure Drop (Initial)", "Δp_initial", results.get('dp_kpa_initial', 'N/A'), "kPa"]
            ]
            
            for r, row_data in enumerate(output_data, start=output_row+1):
                for c, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell_value, float):
                        cell.value = round(cell_value, 4)
                    else:
                        cell.value = cell_value
                    cell.border = border
        
        # Adjust column widths
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if hasattr(cell, 'value') and cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 20)  # Cap at 20 for readability
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(filename)
        print(f"Results saved to {filename}")


    def read_inputs_from_excel(self, input_filename="cooling_circuit_inputs.xlsm"):
        """
        Read inputs from Excel file instead of terminal input.
        
        Args:
            input_filename: Name of the Excel file containing inputs
            
        Returns:
            Dictionary with input parameters
        """
        try:
            # Read the Excel file
            df = pd.read_excel(input_filename, sheet_name="Inputs")
            
            # Convert to dictionary - assuming two columns: Parameter and Value
            inputs = {}
            
            # Check if the Excel has 'Parameter' and 'Value' columns
            if 'Parameter' in df.columns and 'Value' in df.columns:
                for index, row in df.iterrows():
                    param = row['Parameter']
                    value = row['Value']
                    inputs[param] = value
            else:
                # Alternative format: first column is parameter, second is value
                df.columns = ['Parameter', 'Value']  # Rename columns
                for index, row in df.iterrows():
                    if pd.notna(row['Parameter']) and pd.notna(row['Value']):
                        inputs[row['Parameter']] = row['Value']
            
            # Validate and convert data types
            validated_inputs = self.validate_inputs(inputs)
            
            return validated_inputs
            
        except FileNotFoundError:
            print(f"Error: Input file '{input_filename}' not found.")
            print("Please create an Excel file with the following format:")
            self.create_input_template(input_filename)
            return None
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None

    def validate_inputs(self, inputs):
        """
        Validate and convert input data types.
        """
        validated = {}
        
        # Define expected data types and validation rules
        validation_rules = {
            'NK': {'type': int, 'min': 1, 'max': 9},
            'NB': {'type': int, 'min': 2, 'max': 9},
            'HSP': {'type': float, 'min': 0},
            'ASP': {'type': float, 'min': 0},
            'min_bending_radius': {'type': float, 'min': 0},
            'I': {'type': float, 'min': 0},
            'NL': {'type': int, 'min': 1},
            'BCU': {'type': float, 'min': 0},
            'HCU': {'type': float, 'min': 0},
            'TCU': {'type': float, 'min': 0},
            'RICU': {'type': float, 'min': 0},
            'RYCU': {'type': float, 'min': 0},
            'Tin': {'type': float},
            'Tmax': {'type': float},
            'TCISO': {'type': float, 'allowed_values': [2.4, 2.8, 3.2]}
        }
        
        for param, rules in validation_rules.items():
            if param not in inputs:
                raise ValueError(f"Missing required parameter: {param}")
            
            value = inputs[param]
            
            # Convert to appropriate type
            try:
                if rules['type'] == int:
                    validated[param] = int(value)
                elif rules['type'] == float:
                    validated[param] = float(value)
            except (ValueError, TypeError):
                raise ValueError(f"Invalid value for {param}: {value}")
            
            # Validate ranges
            if 'min' in rules and validated[param] < rules['min']:
                raise ValueError(f"{param} must be >= {rules['min']}")
            if 'max' in rules and validated[param] > rules['max']:
                raise ValueError(f"{param} must be <= {rules['max']}")
            if 'allowed_values' in rules and validated[param] not in rules['allowed_values']:
                raise ValueError(f"{param} must be one of {rules['allowed_values']}")
        
        return validated

    def create_input_template(self, filename="cooling_circuit_inputs.xlsm"):
            """
            Create an Excel template file with all required input parameters.
            """
            # Define input parameters with descriptions and example values
            input_data = {
                'Parameter': [
                    'NK', 'NB', 'HSP', 'ASP', 'min_bending_radius', 'I', 'NL',
                    'BCU', 'HCU', 'TCU', 'RICU', 'RYCU', 'Tin', 'Tmax', 'TCISO'
                ],
                'Description': [
                    'Number of cooling circuits (1-9)',
                    'Cooling layer (2-9)',
                    'Iron height (mm)',
                    'Iron length (mm)',
                    'Min bending radius (mm)',
                    'Current (A)',
                    'Number of windings per layer',
                    'Copper profile breadth (mm)',
                    'Copper profile height (mm)',
                    'Thickness (mm)',
                    'Inner diameter (mm)',
                    'Outer diameter (mm)',
                    'Inlet temperature (°C)',
                    'Maximum temperature (°C)',
                    'TCISO value (2.4, 2.8, or 3.2)'
                ],
                'Value': [
                    1, 2, 100.0, 150.0, 5.0, 1000.0, 10,
                    15.0, 8.0, 2.0, 3.0, 6.0, 40.0, 80.0, 2.4
                ]
            }
            
            df = pd.DataFrame(input_data)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Inputs', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Inputs']
                
                # Format headers
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Adjust column widths
                for col in range(1, len(df.columns) + 1):
                    max_length = max(len(str(df.iloc[:, col-1].name)), 
                                df.iloc[:, col-1].astype(str).str.len().max())
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2
            
            print(f"Input template created: {filename}")
            print("Please fill in the 'Value' column with your actual values and re-run the script.")

if __name__ == "__main__":
    main()